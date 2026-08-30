from __future__ import annotations

import csv
import argparse
import base64
import hashlib
import json
import math
import os
import re
import sqlite3
import urllib.request
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from dotenv import load_dotenv

ROOT = Path(__file__).parent
load_dotenv(ROOT / ".env")
BUNDLE = ROOT / "candidate_bundle"
DB_PATH = ROOT / "data" / "meridian.db"
OUTPUTS = ROOT / "outputs"
AUDIT_DIR = ROOT / "audit"

HUBS = {
    "Delhi": (28.6139, 77.2090),
    "Gurgaon": (28.4595, 77.0266),
    "Jaipur": (26.9124, 75.7873),
    "Ambala": (30.3782, 76.7767),
    "Chandigarh": (30.7333, 76.7794),
    "Ludhiana": (30.9010, 75.8573),
    "Kanpur": (26.4499, 80.3319),
    "Lucknow": (26.8467, 80.9462),
    "Rudrapur": (28.9875, 79.4141),
}

# Named places are evidence-backed demo anchors, not geofences or live telemetry.
PLACES = {
    **HUBS,
    "Faridabad": (28.4089, 77.3178),
    "Noida": (28.5355, 77.3910),
    "Nainital": (29.3919, 79.4542),
    "Patna": (25.5941, 85.1376),
    "Varanasi": (25.3176, 82.9739),
}

KOBOYO_TRUCK_ICON = "https://koboyo.com/icons/svg/cartoon-truck.svg"
KOBOYO_WARNING_ICON = "https://koboyo.com/icons/svg/warning-sign-for-road.svg"


@lru_cache(maxsize=4)
def _sized_svg_data_url(url: str, width: int, height: int) -> str:
    """Give a dimensionless Koboyo SVG natural dimensions for Deck.gl's bitmap loader."""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Meridian-Hackathon/0.2"})
        with urllib.request.urlopen(req, timeout=5) as response:
            svg = response.read().decode("utf-8")
        svg = re.sub(r"<svg\b", f'<svg width="{width}" height="{height}"', svg, count=1)
        return "data:image/svg+xml;base64," + base64.b64encode(svg.encode()).decode()
    except (OSError, UnicodeDecodeError):
        return url

RULES = [
    ("RULE-DELHI-WINTER", "Delhi winter: BS6 only", "October to February, any route touching Delhi, Gurgaon, Faridabad or Noida requires a BS6 vehicle.", "dispatcher_interview.txt", "route", "critical"),
    ("RULE-HILLS-WINTER", "Winter hill eligibility", "November to February, Rudrapur and Nainital-side routes require an engine heater and no brake work in the previous 30 days.", "dispatcher_interview.txt", "route", "critical"),
    ("RULE-SHAKTI-SLA", "Shakti planning SLA", "Plan Shakti Cement loads to a 36-hour internal SLA even though the paper contract says 48 hours.", "dispatcher_interview.txt; emails/thread_01_shakti_sla.txt", "client", "warning"),
    ("RULE-VERTEX-GATE", "Vertex Ludhiana gate", "Vertex Ludhiana accepts 08:00–18:00. Hold after-hours arrivals, notify the prior evening, and mark scheduled morning delivery—not failed delivery.", "dispatcher_interview.txt; emails/thread_09_vertex_gate.txt", "client", "warning"),
    ("RULE-APEX-ROTATE", "Apex vehicle rotation", "After any incident on an Apex run, rotate the vehicle for the next Apex dispatch.", "dispatcher_interview.txt; emails/thread_13_apex_rotation.txt", "client", "warning"),
    ("RULE-ORION", "Orion pharma constraints", "Never leave an Orion load unrefrigerated overnight and use a vehicle from model year 2020 or later.", "dispatcher_interview.txt; emails/thread_17_orion_age.txt", "client", "critical"),
    ("RULE-MONSOON", "Eastern monsoon ETA", "July to September, routes east of Lucknow require at least 20% added ETA and no standard SLA promise.", "dispatcher_interview.txt; emails/thread_23_internal_monsoon.txt", "route", "warning"),
    ("RULE-BREAKDOWN", "Breakdown replacement", "Within 50 km of origin, the origin hub sends replacement. Beyond 50 km, use the nearest hub with an eligible vehicle.", "dispatcher_interview.txt", "response", "critical"),
    ("RULE-SERVICE", "Overdue service grounding", "A vehicle more than 30 days past due service is grounded. Missing due-service data is UNKNOWN, never PASS.", "dispatcher_interview.txt", "safety", "critical"),
    ("RULE-JUGAAD", "Temporary repair clock", "A temporary or jugaad repair needs permanent repair within seven days; until repaired, keep the vehicle in its home region.", "dispatcher_interview.txt; emails/thread_25_internal_jugaad.txt", "maintenance", "critical"),
    ("RULE-NIGHT", "New driver night restriction", "A driver with less than six months tenure cannot run solo at night.", "dispatcher_interview.txt; emails/thread_24_internal_nightroster.txt", "driver", "critical"),
]


def connect(path: Path = DB_PATH) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn


def norm_reg(value: str | None) -> str:
    return re.sub(r"[^A-Z0-9]", "", (value or "").upper())


def redact(text: str) -> str:
    text = re.sub(r"(?:\+?91[\s-]?)?[6-9]\d{4}[\s-]?\d{5}", "[PHONE REDACTED]", text)
    text = re.sub(r"\b\d{4}[\s-]\d{4}[\s-]\d{4}\b", "[AADHAAR REDACTED]", text)
    return text


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS source (
          id TEXT PRIMARY KEY, kind TEXT NOT NULL, path TEXT NOT NULL UNIQUE,
          fingerprint TEXT NOT NULL, ingested_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS vehicle (
          registration TEXT PRIMARY KEY, vehicle_id TEXT, model TEXT, year INTEGER,
          bs_stage TEXT, engine_heater TEXT, home_hub TEXT, capacity REAL, status TEXT,
          source_ref TEXT NOT NULL, raw_variants INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS vehicle_conflict (
          id INTEGER PRIMARY KEY, registration TEXT NOT NULL, field TEXT NOT NULL,
          values_json TEXT NOT NULL, source_ref TEXT NOT NULL,
          UNIQUE(registration, field)
        );
        CREATE TABLE IF NOT EXISTS driver (
          driver_id TEXT PRIMARY KEY, joining_date TEXT, home_hub TEXT, source_ref TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS trip (
          trip_id TEXT PRIMARY KEY, created_at TEXT, origin TEXT, destination TEXT,
          vehicle_reg TEXT, driver_id TEXT, client TEXT, status TEXT, actual_time_min REAL,
          source_ref TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS ticket (
          ticket_id TEXT PRIMARY KEY, created_at TEXT, vehicle_reg TEXT, driver_id TEXT,
          origin_hub TEXT, km_from_origin REAL, destination TEXT, issue TEXT, severity TEXT,
          client TEXT, status TEXT, resolution TEXT, valid INTEGER NOT NULL, source_ref TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS maintenance (
          id INTEGER PRIMARY KEY, date TEXT, vehicle_reg TEXT, odometer_km REAL,
          mechanic TEXT, notes TEXT, is_temporary INTEGER NOT NULL,
          is_brake_work INTEGER NOT NULL, source_ref TEXT NOT NULL,
          UNIQUE(date, vehicle_reg, odometer_km, notes)
        );
        CREATE TABLE IF NOT EXISTS rule (
          rule_id TEXT PRIMARY KEY, title TEXT, body TEXT, source_ref TEXT,
          scope TEXT, severity TEXT, version INTEGER NOT NULL DEFAULT 1, status TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS knowledge (
          id TEXT PRIMARY KEY, title TEXT, body TEXT, kind TEXT, language TEXT,
          entity_ref TEXT, location TEXT, valid_until TEXT, source_ref TEXT,
          approved_at TEXT, approved_by TEXT
        );
        CREATE TABLE IF NOT EXISTS proposal (
          id TEXT PRIMARY KEY, created_at TEXT NOT NULL, reporter TEXT NOT NULL,
          transcript TEXT NOT NULL, redacted_text TEXT NOT NULL, kind TEXT NOT NULL,
          language TEXT NOT NULL, entity_ref TEXT, location TEXT, valid_until TEXT,
          status TEXT NOT NULL, source_ref TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS agent_run (
          id TEXT PRIMARY KEY, started_at TEXT NOT NULL, finished_at TEXT,
          provider TEXT NOT NULL, task TEXT NOT NULL, source_ref TEXT,
          status TEXT NOT NULL, trace_json TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS source_upload (
          fingerprint TEXT PRIMARY KEY, name TEXT NOT NULL, media_type TEXT,
          redacted_text TEXT NOT NULL, status TEXT NOT NULL,
          processed_at TEXT NOT NULL, agent_run_id TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS context_event (
          id TEXT PRIMARY KEY, at TEXT NOT NULL, channel TEXT NOT NULL,
          actor_ref TEXT NOT NULL, event_type TEXT NOT NULL, redacted_text TEXT NOT NULL,
          disposition TEXT NOT NULL, reasoning TEXT NOT NULL, source_ref TEXT NOT NULL,
          agent_run_id TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS audit_event (
          id TEXT PRIMARY KEY, at TEXT NOT NULL, actor TEXT NOT NULL,
          action TEXT NOT NULL, object_type TEXT NOT NULL, object_id TEXT NOT NULL,
          details TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS comm_approval (
          ticket_id TEXT PRIMARY KEY, approved_by TEXT NOT NULL, sent_at TEXT NOT NULL
        );
        CREATE VIRTUAL TABLE IF NOT EXISTS context_fts USING fts5(
          title, body, source_ref, object_type, object_id UNINDEXED
        );
        """
    )
    # Existing demo databases migrate in place; rebuild is not required.
    existing = {r[1] for r in conn.execute("PRAGMA table_info(proposal)")}
    for name, declaration in {
        "confidence": "REAL NOT NULL DEFAULT 0.5",
        "reasoning": "TEXT NOT NULL DEFAULT ''",
        "connections_json": "TEXT NOT NULL DEFAULT '[]'",
        "agent_name": "TEXT NOT NULL DEFAULT 'rules-fallback'",
        "risk": "TEXT NOT NULL DEFAULT 'medium'",
    }.items():
        if name not in existing:
            conn.execute(f"ALTER TABLE proposal ADD COLUMN {name} {declaration}")


def _source(conn: sqlite3.Connection, path: Path, kind: str) -> str:
    rel = str(path.relative_to(BUNDLE))
    sid = "SRC-" + hashlib.sha1(rel.encode()).hexdigest()[:12]
    conn.execute(
        "INSERT OR REPLACE INTO source VALUES (?, ?, ?, ?, ?)",
        (sid, kind, rel, sha(path), datetime.now().isoformat(timespec="seconds")),
    )
    return rel


def rebuild(path: Path = DB_PATH) -> sqlite3.Connection:
    if path.exists():
        path.unlink()
    conn = connect(path)
    _schema(conn)
    _ingest_fleet(conn)
    _ingest_drivers(conn)
    _ingest_trips(conn)
    _ingest_tickets(conn)
    _ingest_maintenance(conn)
    _ingest_text(conn)
    for rule in RULES:
        conn.execute("INSERT INTO rule VALUES (?, ?, ?, ?, ?, ?, 1, 'approved')", rule)
    _refresh_fts(conn)
    _audit(conn, "system", "context.rebuilt", "database", "meridian", "All supplied sources ingested")
    conn.commit()
    return conn


def ensure_db() -> sqlite3.Connection:
    conn = connect()
    _schema(conn)
    if not conn.execute("SELECT 1 FROM source LIMIT 1").fetchone():
        conn.close()
        conn = rebuild()
        run_pipeline(conn)
        return conn
    run_pipeline(conn)
    return conn


def _ingest_fleet(conn: sqlite3.Connection) -> None:
    path = BUNDLE / "fleet_master.csv"
    ref = _source(conn, path, "csv")
    groups: dict[str, list[dict]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            groups[norm_reg(row["registration_number"])].append(row)
    for reg, rows in groups.items():
        # ponytail: prefer the identified row; conflicts stay explicit instead of inventing precedence.
        chosen = next((r for r in rows if r["vehicle_id"]), rows[0])
        vals = lambda field: sorted({r[field].strip() for r in rows if r.get(field, "").strip()})
        conn.execute(
            "INSERT INTO vehicle VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (reg, chosen["vehicle_id"] or None, chosen["model"], int(chosen["year"]), chosen["bs_stage"],
             chosen["engine_heater"] or None, chosen["home_hub"], float(chosen["capacity_tonnes"]) if chosen["capacity_tonnes"] else None,
             chosen["status"], ref, len(rows)),
        )
        for field in ("year", "engine_heater", "capacity_tonnes"):
            values = vals(field)
            if len(values) > 1:
                conn.execute("INSERT INTO vehicle_conflict(registration,field,values_json,source_ref) VALUES (?,?,?,?)",
                             (reg, field, json.dumps(values), ref))


def _ingest_drivers(conn: sqlite3.Connection) -> None:
    path = BUNDLE / "drivers_roster.csv"
    ref = _source(conn, path, "csv")
    with path.open(newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            conn.execute("INSERT INTO driver VALUES (?,?,?,?)", (r["driver_id"], r["joining_date"], r["home_hub"], ref))


def _ingest_trips(conn: sqlite3.Connection) -> None:
    path = BUNDLE / "meridian_trips.csv"
    ref = _source(conn, path, "csv")
    with path.open(newline="", encoding="utf-8-sig") as f:
        rows = ((r["trip_id"], r["created_at"], r["origin_name"], r["dest_name"], norm_reg(r["vehicle_reg"]),
                 r["driver_id"], r["client"], r["status"], float(r["actual_time_min"]), ref) for r in csv.DictReader(f))
        conn.executemany("INSERT INTO trip VALUES (?,?,?,?,?,?,?,?,?,?)", rows)


def _ingest_tickets(conn: sqlite3.Connection) -> None:
    path = BUNDLE / "tickets.json"
    ref = _source(conn, path, "json")
    seen: set[str] = set()
    for r in json.loads(path.read_text()):
        tid = str(r.get("ticket_id") or "")
        if tid in seen:
            continue
        seen.add(tid)
        valid = int(bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}", str(r.get("created_at") or ""))
                         and norm_reg(r.get("vehicle")) and r.get("driver_id") and r.get("issue")))
        conn.execute(
            "INSERT INTO ticket VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (tid, r.get("created_at"), norm_reg(r.get("vehicle")), r.get("driver_id"), r.get("origin_hub"),
             r.get("km_from_origin_hub"), r.get("destination"), r.get("issue"), r.get("severity"),
             r.get("client"), r.get("status"), redact(str(r.get("resolution_note") or "")), valid, ref),
        )


def _ingest_maintenance(conn: sqlite3.Connection) -> None:
    path = BUNDLE / "maintenance_log.xlsx"
    ref = _source(conn, path, "xlsx")
    ws = load_workbook(path, read_only=True, data_only=True).active
    for row in ws.iter_rows(min_row=2, values_only=True):
        dt, vehicle, odo, mechanic, notes = row
        note = redact(str(notes or ""))
        low = note.lower()
        conn.execute(
            "INSERT OR IGNORE INTO maintenance(date,vehicle_reg,odometer_km,mechanic,notes,is_temporary,is_brake_work,source_ref) VALUES (?,?,?,?,?,?,?,?)",
            (str(dt)[:10], norm_reg(str(vehicle)), odo, mechanic, note,
             int(any(k in low for k in ("jugaad", "temporary", "permanent fix"))),
             int("brake" in low), ref),
        )


def _ingest_text(conn: sqlite3.Connection) -> None:
    files = [BUNDLE / "dispatcher_interview.txt", *sorted((BUNDLE / "emails").glob("*.txt"))]
    for path in files:
        _source(conn, path, "interview" if "interview" in path.name else "email")


def _refresh_fts(conn: sqlite3.Connection) -> None:
    conn.execute("DELETE FROM context_fts")
    for row in conn.execute("SELECT rule_id,title,body,source_ref FROM rule"):
        conn.execute("INSERT INTO context_fts VALUES (?,?,?,?,?)", (row["title"], row["body"], row["source_ref"], "rule", row["rule_id"]))
    for path in [BUNDLE / "dispatcher_interview.txt", *sorted((BUNDLE / "emails").glob("*.txt"))]:
        rel = str(path.relative_to(BUNDLE))
        conn.execute("INSERT INTO context_fts VALUES (?,?,?,?,?)", (path.stem.replace("_", " ").title(), redact(path.read_text(errors="replace")), rel, "source", rel))
    for row in conn.execute("SELECT id,title,body,source_ref FROM knowledge"):
        conn.execute("INSERT INTO context_fts VALUES (?,?,?,?,?)", (row["title"], row["body"], row["source_ref"], "knowledge", row["id"]))


def _audit(conn: sqlite3.Connection, actor: str, action: str, object_type: str, object_id: str, details: str) -> None:
    conn.execute("INSERT INTO audit_event VALUES (?,?,?,?,?,?,?)",
                 (str(uuid.uuid4()), datetime.now().isoformat(timespec="seconds"), actor, action, object_type, object_id, details))


def stats(conn: sqlite3.Connection) -> dict[str, int]:
    scalar = lambda q: int(conn.execute(q).fetchone()[0])
    return {
        "vehicles": scalar("SELECT count(*) FROM vehicle"),
        "trips": scalar("SELECT count(*) FROM trip"),
        "drivers": scalar("SELECT count(*) FROM driver"),
        "rules": scalar("SELECT count(*) FROM rule WHERE status='approved'"),
        "conflicts": scalar("SELECT count(*) FROM vehicle_conflict"),
        "invalid_tickets": scalar("SELECT count(*) FROM ticket WHERE valid=0"),
        "temporary_repairs": scalar("SELECT count(*) FROM maintenance WHERE is_temporary=1"),
        "pending": scalar("SELECT count(*) FROM proposal WHERE status='pending'"),
    }


def data_quality(conn: sqlite3.Connection) -> list[dict]:
    raw_fleet = sum(1 for _ in csv.DictReader((BUNDLE / "fleet_master.csv").open()))
    canonical = conn.execute("SELECT count(*) FROM vehicle").fetchone()[0]
    raw_tickets = len(json.loads((BUNDLE / "tickets.json").read_text()))
    unique_tickets = conn.execute("SELECT count(*) FROM ticket").fetchone()[0]
    regressions = 0
    by_vehicle: dict[str, list[tuple[str, float]]] = defaultdict(list)
    for r in conn.execute("SELECT date,vehicle_reg,odometer_km FROM maintenance ORDER BY vehicle_reg,date"):
        if r["odometer_km"] is not None:
            by_vehicle[r["vehicle_reg"]].append((r["date"], r["odometer_km"]))
    for points in by_vehicle.values():
        regressions += sum(b[1] < a[1] for a, b in zip(points, points[1:]))
    return [
        {"severity": "warning", "title": f"{raw_fleet - canonical} duplicate fleet rows merged", "detail": f"{raw_fleet} source rows → {canonical} canonical registrations", "source": "fleet_master.csv"},
        {"severity": "critical", "title": "Model-year conflicts block automatic age decisions", "detail": "CH42HD4155, RJ34GX8933 and UP40IM3144 disagree inside the fleet source; RJ43DD3546 also conflicts with an email claim.", "source": "fleet_master.csv + email"},
        {"severity": "warning", "title": f"{raw_tickets - unique_tickets} duplicate ticket rows ignored", "detail": f"{raw_tickets} source rows → {unique_tickets} unique ticket IDs", "source": "tickets.json"},
        {"severity": "critical", "title": "2 tickets quarantined", "detail": "TKT-9101 and TKT-9102 lack required identity/date/location evidence.", "source": "tickets.json"},
        {"severity": "warning", "title": f"{regressions} odometer regression events", "detail": "Maintenance mileage cannot be treated as a monotonic current odometer without reconciliation.", "source": "maintenance_log.xlsx"},
        {"severity": "info", "title": "Trip history is not live telemetry", "detail": "10,000 trips are from Sep–Oct 2018. Map locations show home assignment or historical routes—not current vehicle positions.", "source": "meridian_trips.csv"},
    ]


def hub_rows(conn: sqlite3.Connection) -> list[dict]:
    counts = {r["home_hub"]: r["n"] for r in conn.execute("SELECT home_hub,count(*) n FROM vehicle GROUP BY home_hub")}
    incidents = {r["origin_hub"]: r["n"] for r in conn.execute("SELECT origin_hub,count(*) n FROM ticket WHERE status!='CLOSED' AND valid=1 GROUP BY origin_hub")}
    rows = []
    for name, (lat, lon) in HUBS.items():
        rows.append({"hub": name, "lat": lat, "lon": lon, "vehicles": counts.get(name, 0), "incidents": incidents.get(name, 0), "position_type": "home assignment"})
    return rows


def truck_rows(conn: sqlite3.Connection) -> list[dict]:
    """One visible truck glyph per canonical home assignment; never a live position."""
    rows: list[dict] = []
    for vehicle in conn.execute("SELECT registration,model,year,bs_stage,home_hub,status FROM vehicle ORDER BY home_hub,registration"):
        if vehicle["home_hub"] not in HUBS:
            continue
        lat, lon = HUBS[vehicle["home_hub"]]
        digest = int(hashlib.sha1(vehicle["registration"].encode()).hexdigest()[:8], 16)
        angle = (digest % 360) * math.pi / 180
        radius = .025 + ((digest // 360) % 6) * .012
        rows.append({
            **dict(vehicle), "lat": lat + math.sin(angle) * radius,
            "lon": lon + math.cos(angle) * radius,
            "icon": {"url": _sized_svg_data_url(KOBOYO_TRUCK_ICON, 259, 259), "width": 259, "height": 259, "anchorY": 259},
            "evidence": "fleet-master home assignment; not parked/live telemetry",
        })
    return rows


@lru_cache(maxsize=48)
def road_route(origin: str, destination: str) -> dict:
    """Return OSRM road geometry, with an explicit approximate fallback."""
    if origin not in PLACES or destination not in PLACES:
        raise ValueError("Unknown route endpoint")
    a_lat, a_lon = PLACES[origin]
    b_lat, b_lon = PLACES[destination]
    url = (
        "https://router.project-osrm.org/route/v1/driving/"
        f"{a_lon},{a_lat};{b_lon},{b_lat}?overview=full&geometries=geojson"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "Meridian-Hackathon/0.2 (route-intelligence-demo)"})
    try:
        with urllib.request.urlopen(req, timeout=7) as response:
            route = json.loads(response.read())["routes"][0]
        return {
            "path": route["geometry"]["coordinates"],
            "distance_km": round(route["distance"] / 1000, 1),
            "duration_hr": round(route["duration"] / 3600, 1),
            "geometry_source": "OSRM road route",
            "is_approximate": False,
        }
    except (OSError, KeyError, IndexError, json.JSONDecodeError):
        # A curved interpolation keeps the UI useful offline without claiming a road route.
        path = []
        for i in range(21):
            t = i / 20
            bend = math.sin(math.pi * t) * .18
            path.append([a_lon + (b_lon - a_lon) * t + bend, a_lat + (b_lat - a_lat) * t])
        straight_km = 111 * math.sqrt((b_lat - a_lat) ** 2 + ((b_lon - a_lon) * math.cos(math.radians((a_lat + b_lat) / 2))) ** 2)
        return {
            "path": path, "distance_km": round(straight_km, 1), "duration_hr": None,
            "geometry_source": "approximate endpoint connection (OSRM unavailable)", "is_approximate": True,
        }


def _point_on_path(path: list[list[float]], fraction: float) -> list[float]:
    if not path:
        return [0, 0]
    idx = min(len(path) - 1, max(0, round(fraction * (len(path) - 1))))
    return path[idx]


def route_intelligence(conn: sqlite3.Connection, origin: str, destination: str, client: str, travel_on: date | str) -> dict:
    """Compile evidence, rules, historical incidents and bounded fleet checks for a route."""
    when = date.fromisoformat(travel_on) if isinstance(travel_on, str) else travel_on
    route = road_route(origin, destination)
    cities = {origin, destination}
    lower_client = client.lower()
    precautions: list[dict] = []

    def add(rule_id: str, status: str, why: str) -> None:
        row = conn.execute("SELECT * FROM rule WHERE rule_id=?", (rule_id,)).fetchone()
        if row:
            precautions.append({**dict(row), "status": status, "why_now": why})

    if when.month in (10, 11, 12, 1, 2) and cities & {"Delhi", "Gurgaon", "Faridabad", "Noida"}:
        add("RULE-DELHI-WINTER", "BLOCKING", "Selected date is inside the Delhi-NCR winter window.")
    if when.month in (11, 12, 1, 2) and cities & {"Rudrapur", "Nainital"}:
        add("RULE-HILLS-WINTER", "BLOCKING", "Selected route/date touches the winter hill corridor.")
    if "shakti" in lower_client:
        add("RULE-SHAKTI-SLA", "PLAN", "Client selected: Shakti Cement.")
    if "vertex" in lower_client and "Ludhiana" in cities:
        add("RULE-VERTEX-GATE", "PLAN", "Vertex delivery touches Ludhiana.")
    if "apex" in lower_client:
        add("RULE-APEX-ROTATE", "CHECK", "Apex history must be checked before assignment.")
    if "orion" in lower_client:
        add("RULE-ORION", "BLOCKING", "Orion load constraints apply.")
    if when.month in (7, 8, 9) and (destination in {"Patna", "Varanasi"} or origin in {"Patna", "Varanasi"}):
        add("RULE-MONSOON", "PLAN", "Selected route/date is in the eastern monsoon window.")
    add("RULE-SERVICE", "DATA GAP", "Current service-due state is absent, so no vehicle can be marked fully eligible.")

    incidents: list[dict] = []
    for row in conn.execute(
        "SELECT ticket_id,km_from_origin,issue,severity,status,source_ref FROM ticket WHERE valid=1 AND origin_hub=? AND destination=? ORDER BY created_at DESC",
        (origin, destination),
    ):
        fraction = min(.96, max(.04, (row["km_from_origin"] or 0) / max(route["distance_km"], 1)))
        lon, lat = _point_on_path(route["path"], fraction)
        incidents.append({**dict(row), "lat": lat, "lon": lon, "position_basis": "historical km-from-origin projected onto selected route"})

    candidates: list[dict] = []
    has_delhi = any(p["rule_id"] == "RULE-DELHI-WINTER" for p in precautions)
    has_hills = any(p["rule_id"] == "RULE-HILLS-WINTER" for p in precautions)
    has_orion = "orion" in lower_client
    for vehicle in conn.execute("SELECT * FROM vehicle WHERE home_hub=? ORDER BY registration", (origin,)):
        checks, blocked = [], False
        conflict_fields = [row["field"] for row in conn.execute("SELECT field FROM vehicle_conflict WHERE registration=?", (vehicle["registration"],))]
        if conflict_fields:
            checks.append("Conflicting canonical fields: " + ", ".join(conflict_fields))
            blocked = True
        if has_delhi:
            ok = vehicle["bs_stage"] == "BS6"
            checks.append(f"Delhi winter: {'PASS' if ok else 'FAIL'} ({vehicle['bs_stage']})")
            blocked |= not ok
        if has_hills:
            ok = (vehicle["engine_heater"] or "").strip().lower() in {"yes", "true", "1", "y"}
            checks.append(f"Engine heater: {'PASS' if ok else 'FAIL/UNKNOWN'}")
            blocked |= not ok
            cutoff = (when - timedelta(days=30)).isoformat()
            recent_brakes = conn.execute("SELECT count(*) FROM maintenance WHERE vehicle_reg=? AND is_brake_work=1 AND date BETWEEN ? AND ?", (vehicle["registration"], cutoff, when.isoformat())).fetchone()[0]
            checks.append(f"Brake work prior 30d: {'FAIL' if recent_brakes else 'no recorded event'}")
            blocked |= bool(recent_brakes)
        if has_orion:
            conflicts = conn.execute("SELECT 1 FROM vehicle_conflict WHERE registration=? AND field='year'", (vehicle["registration"],)).fetchone()
            ok = vehicle["year"] >= 2020 and not conflicts
            checks.append(f"Orion 2020+: {'PASS' if ok else 'FAIL/CONFLICT'} ({vehicle['year']})")
            blocked |= not ok
        checks += ["Current availability: LIVE FEED MISSING", "Service due: FIELD ABSENT"]
        candidates.append({
            "registration": vehicle["registration"], "model": vehicle["model"], "year": vehicle["year"],
            "bs_stage": vehicle["bs_stage"], "assessment": "STATIC BLOCK" if blocked else "CONDITIONAL",
            "checks": checks, "note": "Never a dispatch PASS until live availability and service state are connected.",
        })

    midpoint = _point_on_path(route["path"], .5)
    for i, item in enumerate(precautions):
        point = _point_on_path(route["path"], min(.88, .15 + i * .11))
        item.update({
            "lon": point[0], "lat": point[1],
            "icon": {"url": _sized_svg_data_url(KOBOYO_WARNING_ICON, 185, 174), "width": 185, "height": 174, "anchorY": 174},
        })
    trip_span = conn.execute("SELECT substr(min(created_at),1,10),substr(max(created_at),1,10) FROM trip").fetchone()
    maintenance_count = conn.execute("SELECT count(*) FROM maintenance").fetchone()[0]
    origin_conflicts = [dict(row) for row in conn.execute(
        "SELECT vc.registration,vc.field,vc.source_ref FROM vehicle_conflict vc JOIN vehicle v ON v.registration=vc.registration WHERE v.home_hub=? ORDER BY vc.registration",
        (origin,),
    )]
    return {
        **route, "origin": origin, "destination": destination, "client": client, "travel_on": when.isoformat(),
        "precautions": precautions, "incidents": incidents, "candidates": candidates,
        "midpoint": {"lon": midpoint[0], "lat": midpoint[1]},
        "unknowns": ["current vehicle positions", "parked count now", "live availability", "verified service-due state", "live road/weather conditions"],
        "uncertainty_groups": [
            {
                "label": "Live feeds not connected",
                "items": ["current vehicle positions", "parked count now", "dispatch availability", "live road and weather"],
                "effect": "Final dispatch state cannot be claimed.",
            },
            {
                "label": "Required field absent",
                "items": ["verified service-due date/state"],
                "effect": f"{maintenance_count} maintenance events exist, but none supplies a service-due field.",
            },
            {
                "label": "Historical, not current",
                "items": [f"trip history covers {trip_span[0]} to {trip_span[1]}"],
                "effect": "Useful for patterns and precedent, not live conditions.",
            },
        ],
        "origin_conflicts": origin_conflicts,
    }


def search(conn: sqlite3.Connection, question: str, limit: int = 6) -> list[dict]:
    tokens = re.findall(r"[A-Za-z0-9]+", question)
    if not tokens:
        return []
    match = " OR ".join(f'"{t}"' for t in tokens[:10])
    try:
        return [dict(r) for r in conn.execute(
            "SELECT title,body,source_ref,object_type,object_id,bm25(context_fts) score FROM context_fts WHERE context_fts MATCH ? ORDER BY score LIMIT ?",
            (match, limit),
        )]
    except sqlite3.OperationalError:
        return []


def answer(conn: sqlite3.Connection, question: str) -> dict:
    q = question.lower()
    hindi = any(k in q for k in ("hai", "kya", "ke liye", "gaadi", "bhej", "karna", "sahi"))
    reg_match = re.search(r"\b(?:[a-z]{2}[\s-]?\d{1,2}[\s-]?[a-z]{1,2}[\s-]?\d{4})\b", question, re.I)
    citations: list[str] = []
    unknowns: list[str] = []
    headline = "यह फैसला उपलब्ध संदर्भ से नहीं निकलता।" if hindi else "The available context is not enough for a safe decision."
    detail = ""
    if reg_match:
        reg = norm_reg(reg_match.group())
        v = conn.execute("SELECT * FROM vehicle WHERE registration=?", (reg,)).fetchone()
        if v:
            citations.append("fleet_master.csv")
            if "orion" in q:
                citations += ["dispatcher_interview.txt", "emails/thread_17_orion_age.txt"]
                eligible = v["year"] >= 2020
                headline = (f"नहीं—{reg} Orion के लिए age rule में FAIL है।" if hindi and not eligible else
                            f"हाँ—{reg} Orion के 2020+ age rule में PASS है।" if hindi else
                            f"{'PASS' if eligible else 'FAIL'} — {reg} is {'new enough' if eligible else 'too old'} for Orion.")
                detail = f"Fleet master says model year {v['year']}; Orion requires 2020 or later. BS stage: {v['bs_stage']}; home assignment: {v['home_hub']}."
                if reg == "RJ43DD3546":
                    detail += " An email calls it a brand-new 2021 vehicle, conflicting with the fleet master's 2017; the conservative result remains FAIL pending RC verification."
            elif any(k in q for k in ("delhi", "gurgaon", "noida", "faridabad")):
                citations.append("dispatcher_interview.txt")
                eligible = v["bs_stage"] == "BS6"
                headline = f"{'PASS' if eligible else 'FAIL'} — {reg} is {v['bs_stage']} for the Delhi winter BS6 rule."
                detail = "This restriction applies October through February on routes touching Delhi NCR."
            else:
                headline = f"{reg}: {v['model']}, {v['year']}, {v['bs_stage']}, home assignment {v['home_hub']}."
                detail = "Current location, availability and service-due state are not present in the supplied sources."
                unknowns += ["current GPS/location", "current availability", "verified service-due status"]
    elif "shakti" in q and any(k in q for k in ("sla", "hour", "time", "kitna", "deadline")):
        headline = "Shakti ke liye 36 ghante ka internal SLA plan karo." if hindi else "Plan Shakti loads to a 36-hour internal SLA."
        detail = "The paper contract says 48 hours, but the dispatcher interview and email evidence say operations escalate after 36."
        citations = ["dispatcher_interview.txt", "emails/thread_01_shakti_sla.txt"]
    elif any(k in q for k in ("breakdown", "toot", "kharab")) and any(k in q for k in ("50", "replacement", "badli")):
        headline = "50 km ke andar origin hub replacement bhejega; uske baad nearest eligible hub." if hindi else "Within 50 km, the origin hub sends replacement; beyond 50 km, use the nearest eligible hub."
        detail = "Eligibility still requires route/season compliance, maintenance safety and availability."
        citations = ["dispatcher_interview.txt"]
        unknowns = ["live hub inventory", "current service-due state"]
    else:
        evidence = search(conn, question)
        if evidence:
            top = evidence[0]
            headline = top["title"]
            detail = top["body"][:650].strip()
            citations = list(dict.fromkeys(r["source_ref"] for r in evidence[:3]))
        unknowns.append("No live operational feed is connected; verify current state before dispatch.")
    return {"headline": headline, "detail": detail, "citations": list(dict.fromkeys(citations)), "unknowns": unknowns}


def detect_language(text: str) -> str:
    if re.search(r"[\u0900-\u097F]", text):
        return "Hindi"
    if any(k in text.lower().split() for k in ("hai", "tha", "gaadi", "jugaad", "kal", "aaj", "nahi")):
        return "Hinglish"
    return "English"


def classify(text: str) -> str:
    low = text.lower()
    if any(k in low for k in ("route", "road", "rasta", "raasta", "bridge", "diversion")):
        return "route precaution"
    if any(k in low for k in ("repair", "brake", "engine", "tyre", "leak", "jugaad")):
        return "maintenance observation"
    if any(k in low for k in ("client", "gate", "warehouse", "sla", "delivery")):
        return "client rule"
    return "ground observation"


def propose(
    conn: sqlite3.Connection, reporter: str, transcript: str, location: str = "", valid_until: str = "",
    entity_ref: str = "", *, kind: str | None = None, confidence: float = .5, reasoning: str = "",
    connections: list[str] | None = None, agent_name: str = "rules-fallback", risk: str = "medium",
    source_ref: str = "worker voice/text",
) -> str:
    pid = "PROP-" + uuid.uuid4().hex[:10].upper()
    clean = redact(transcript.strip())
    proposal_kind = kind or classify(clean)
    conn.execute(
        """INSERT INTO proposal(
          id,created_at,reporter,transcript,redacted_text,kind,language,entity_ref,location,valid_until,
          status,source_ref,confidence,reasoning,connections_json,agent_name,risk
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (pid, datetime.now().isoformat(timespec="seconds"), reporter.strip() or "Anonymous worker", clean, clean,
         proposal_kind, detect_language(clean), norm_reg(entity_ref) or None, location or None, valid_until or None,
         "pending", source_ref, max(0, min(1, float(confidence))), reasoning.strip(),
         json.dumps(connections or [], ensure_ascii=False), agent_name, risk),
    )
    _audit(conn, agent_name, "proposal.created", "proposal", pid, f"{proposal_kind} staged from {source_ref}; awaiting human approval")
    conn.commit()
    return pid


def decide_proposal(conn: sqlite3.Connection, proposal_id: str, approve: bool, actor: str = "Operations approver") -> None:
    row = conn.execute("SELECT * FROM proposal WHERE id=? AND status='pending'", (proposal_id,)).fetchone()
    if not row:
        return
    status = "approved" if approve else "rejected"
    conn.execute("UPDATE proposal SET status=? WHERE id=?", (status, proposal_id))
    if approve:
        kid = "KNOW-" + proposal_id.removeprefix("PROP-")
        conn.execute("INSERT INTO knowledge VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                     (kid, row["kind"].title(), row["redacted_text"], row["kind"], row["language"], row["entity_ref"], row["location"],
                      row["valid_until"], row["source_ref"] + "; " + proposal_id, datetime.now().isoformat(timespec="seconds"), actor))
        _refresh_fts(conn)
    _audit(conn, actor, f"proposal.{status}", "proposal", proposal_id, "Human decision recorded")
    conn.commit()


def transcribe_sarvam(audio_bytes: bytes, mime: str = "audio/wav") -> str:
    key = os.getenv("SARVAM_API_KEY")
    if not key:
        raise RuntimeError("SARVAM_API_KEY is not configured")
    boundary = "----Meridian" + uuid.uuid4().hex
    chunks = [
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"model\"\r\n\r\nsaaras:v3\r\n".encode(),
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"language_code\"\r\n\r\nunknown\r\n".encode(),
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; filename=\"worker-note.wav\"\r\nContent-Type: {mime}\r\n\r\n".encode(),
        audio_bytes,
        f"\r\n--{boundary}--\r\n".encode(),
    ]
    req = urllib.request.Request("https://api.sarvam.ai/speech-to-text", data=b"".join(chunks), method="POST",
                                 headers={"api-subscription-key": key, "Content-Type": f"multipart/form-data; boundary={boundary}"})
    with urllib.request.urlopen(req, timeout=60) as response:
        return json.loads(response.read())["transcript"]


def export_audit(conn: sqlite3.Connection) -> str:
    rows = [dict(r) for r in conn.execute("SELECT * FROM audit_event ORDER BY at,id")]
    return "\n".join(json.dumps(r, sort_keys=True, ensure_ascii=False) for r in rows) + "\n"


ALIASES = {
    "ticket_id": ("ticket_id", "ticketId", "id"),
    "created_at": ("created_at", "createdAt", "reported_at", "timestamp"),
    "vehicle": ("vehicle", "vehicle_reg", "registration", "vehicle_registration"),
    "driver_id": ("driver_id", "driverId", "driver"),
    "origin_hub": ("origin_hub", "origin", "source_hub"),
    "km_from_origin_hub": ("km_from_origin_hub", "km_from_origin", "distance_from_origin_km"),
    "destination": ("destination", "dest", "destination_hub"),
    "issue": ("issue", "problem", "description"),
    "severity": ("severity", "priority"),
    "client": ("client", "customer"),
    "status": ("status", "ticket_status"),
    "resolution_note": ("resolution_note", "resolution", "notes"),
}


def _adapt_ticket(raw: dict) -> dict:
    return {field: next((raw.get(k) for k in names if raw.get(k) not in (None, "")), None) for field, names in ALIASES.items()}


def _ticket_errors(ticket: dict) -> list[str]:
    errors = []
    required = ("ticket_id", "created_at", "vehicle", "driver_id", "origin_hub", "km_from_origin_hub", "destination", "issue", "severity", "client")
    errors.extend(f"missing {k}" for k in required if ticket.get(k) in (None, ""))
    try:
        datetime.fromisoformat(str(ticket.get("created_at")))
    except ValueError:
        errors.append("invalid created_at")
    if ticket.get("vehicle") and not norm_reg(str(ticket["vehicle"])):
        errors.append("invalid vehicle")
    return sorted(set(errors))


def _deterministic_id(prefix: str, value: str) -> str:
    return f"{prefix}-" + hashlib.sha256(value.encode()).hexdigest()[:14].upper()


def _jsonl(path: Path, rows: Iterable[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content = "\n".join(json.dumps(row, sort_keys=True, ensure_ascii=False, separators=(",", ":")) for row in rows)
    path.write_text(content + ("\n" if content else ""), encoding="utf-8")


def run_pipeline(conn: sqlite3.Connection, ticket_paths: Iterable[Path] | None = None) -> dict[str, int]:
    paths = list(ticket_paths or [BUNDLE / "tickets.json"])
    canonical: dict[str, dict] = {}
    duplicate_counts: Counter = Counter()
    source_for: dict[str, str] = {}
    for path in paths:
        payload = json.loads(path.read_text())
        records = payload if isinstance(payload, list) else payload.get("tickets", payload.get("records", []))
        for raw in records:
            ticket = _adapt_ticket(raw)
            tid = str(ticket.get("ticket_id") or _deterministic_id("MISSING", json.dumps(raw, sort_keys=True)))
            duplicate_counts[tid] += 1
            canonical.setdefault(tid, ticket)
            source_for.setdefault(tid, path.name)

    work_orders, pending, quarantine, audit = [], [], [], []
    for tid in sorted(canonical):
        t = canonical[tid]
        source = source_for[tid]
        errors = _ticket_errors(t)
        if errors:
            quarantine.append({"ticket_id": tid, "reasons": errors, "source": source})
            audit.append({"step": "quarantined", "ticket_id": tid, "decision": "; ".join(errors), "citations": [source]})
            continue
        reg = norm_reg(str(t["vehicle"]))
        km = float(t["km_from_origin_hub"])
        response = f"Origin hub ({t['origin_hub']}) sends replacement" if km <= 50 else "Nearest hub with an eligible vehicle sends replacement"
        citations = [source, "dispatcher_interview.txt"]
        wo_id = _deterministic_id("WO", tid)
        msg_id = _deterministic_id("MSG", tid)
        work_orders.append({
            "work_order_id": wo_id, "ticket_id": tid, "vehicle_reg": reg,
            "created_at": str(t["created_at"]), "citations": citations,
        })
        body = (f"Meridian update {tid}: a {str(t['severity']).lower()} {t['issue']} incident is recorded "
                f"for vehicle {reg} on the {t['origin_hub']}–{t['destination']} movement. {response}. "
                "This draft awaits operations approval.")
        pending.append({
            "message_id": msg_id, "ticket_id": tid, "recipient": t["client"], "body": redact(body),
            "context": {"severity": t["severity"], "status": t.get("status"), "km_from_origin_hub": km, "response": response},
            "citations": citations, "approval_status": "pending",
        })
        audit.extend([
            {"step": "normalized", "ticket_id": tid, "decision": f"canonicalized; {duplicate_counts[tid]} source row(s)", "citations": [source]},
            {"step": "response_decided", "ticket_id": tid, "decision": response, "rule_id": "RULE-BREAKDOWN", "citations": citations},
            {"step": "outputs_drafted", "ticket_id": tid, "decision": f"{wo_id}; {msg_id}", "citations": citations},
        ])

    approvals = {r["ticket_id"]: dict(r) for r in conn.execute("SELECT * FROM comm_approval")}
    sent = []
    for draft in pending:
        approval = approvals.get(draft["ticket_id"])
        if approval:
            sent.append({"message_id": draft["message_id"], "ticket_id": draft["ticket_id"], "recipient": draft["recipient"],
                         "body": draft["body"].replace(" This draft awaits operations approval.", ""),
                         "approved_by": approval["approved_by"], "sent_at": approval["sent_at"]})

    _jsonl(OUTPUTS / "work_orders.jsonl", work_orders)
    _jsonl(OUTPUTS / "comms_pending.jsonl", pending)
    _jsonl(OUTPUTS / "comms_sent.jsonl", sent)
    _jsonl(OUTPUTS / "quarantine.jsonl", quarantine)
    _jsonl(AUDIT_DIR / "audit.jsonl", audit)
    return {"work_orders": len(work_orders), "pending": len(pending), "sent": len(sent), "quarantine": len(quarantine), "audit_events": len(audit)}


def approve_communication(conn: sqlite3.Connection, ticket_id: str, approved_by: str) -> None:
    conn.execute("INSERT OR IGNORE INTO comm_approval VALUES (?,?,?)", (ticket_id, approved_by, datetime.now().isoformat(timespec="seconds")))
    _audit(conn, approved_by, "communication.approved", "ticket", ticket_id, "Draft moved to comms_sent exactly once")
    conn.commit()
    run_pipeline(conn)


def cli() -> None:
    parser = argparse.ArgumentParser(description="Meridian deterministic context pipeline")
    parser.add_argument("command", choices=["pipeline", "rebuild"])
    parser.add_argument("--tickets", nargs="*", type=Path, help="Ticket JSON files; aliases are normalized")
    args = parser.parse_args()
    conn = rebuild() if args.command == "rebuild" else ensure_db()
    result = run_pipeline(conn, args.tickets or None)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    cli()
